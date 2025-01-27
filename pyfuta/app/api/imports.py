from io import BytesIO
from typing import List
from fastapi import APIRouter, Depends, File, HTTPException
from openpyxl import load_workbook
from pyfuta.app.database import require_session
from pyfuta.app.models.report import Report, ReportField
from pyfuta.app.api.reports import require_fields, require_report
from sqlalchemy import Table
from sqlalchemy.dialects.sqlite import insert
from sqlmodel import SQLModel, Session
from pyfuta.app.database import async_engine

router = APIRouter(prefix="/imports", tags=["Imports"])


@router.post("/excel")
async def import_excel(
    report: Report = Depends(require_report),
    fields: List[ReportField] = Depends(require_fields),
    session: Session = Depends(require_session),
    file: bytes = File(),
):
    workbook = load_workbook(BytesIO(file), read_only=True)
    sheet = workbook.worksheets[0]
    headers = [cell.value for cell in sheet[1]]
    table_headers = [field.linked_field for field in fields if field.linked_field is not None]
    if report.linked_table is None:
        raise HTTPException(status_code=400, detail="Report is not bound to a table")
    if len(headers) != len(table_headers):
        raise HTTPException(status_code=400, detail="Number of columns in the file does not match the number of table fields in the report")

    parsed_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        parsed_rows.append([field.type.parse(cell) for field, cell in zip(fields, row)])

    # we use sqlite dialect for insert on conflict do nothing
    table = Table(report.linked_table, SQLModel.metadata, autoload_with=session.bind)
    statement = insert(table).on_conflict_do_nothing()
    sentences = [dict(zip(table_headers, row)) for row in parsed_rows]

    async with async_engine.begin() as conn:
        await conn.execute(statement, sentences)
        await conn.commit()
