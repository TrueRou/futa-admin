from io import BytesIO
from fastapi import APIRouter, Depends, File, HTTPException
from openpyxl import load_workbook
from pyfuta.app.database import require_session
from pyfuta.app.reports.models import Report, ReportField
from pyfuta.app.reports.routes import require_fields, require_report
from sqlalchemy import Table
from sqlalchemy.dialects.sqlite import insert
from sqlmodel import SQLModel, Session
from pyfuta.app.database import async_engine

router = APIRouter(prefix="/imports", tags=["imports"])


@router.post("/excel")
async def import_excel(
    report: Report = Depends(require_report),
    fields: list[ReportField] = Depends(require_fields),
    session: Session = Depends(require_session),
    file: bytes = File(),
):
    workbook = load_workbook(BytesIO(file), read_only=True)
    sheet = workbook.worksheets[0]
    headers = [cell.value for cell in sheet[1]]
    table_headers = [field.field_name for field in fields if field.field_name is not None]
    if report.table_name is None:
        raise HTTPException(status_code=400, detail="Report is not bound to a table")
    if len(headers) != len(table_headers):
        raise HTTPException(status_code=400, detail="Number of columns in the file does not match the number of table fields in the report")

    # we use sqlite dialect for insert on conflict do nothing
    table = Table(report.table_name, SQLModel.metadata, autoload_with=session.bind)
    statement = insert(table).on_conflict_do_nothing()
    sentences = [dict(zip(table_headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
    async with async_engine.begin() as conn:
        await conn.execute(statement, sentences)
        await conn.commit()
